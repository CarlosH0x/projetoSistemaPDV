import tkinter as tk
from tkinter import messagebox, ttk
import sqlite3
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Lista para armazenar os produtos da venda atual
venda_atual = []

def conectar():
    conn = sqlite3.connect('sistema_vendas.db')
    cursor = conn.cursor()
    
    # Criação da tabela de 'vendas' se ela não existir
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS vendas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produto TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            preco REAL NOT NULL,
            data DATE DEFAULT (datetime('now', 'localtime'))
        )
    ''')
    
    conn.commit()
    return conn

def registrar_venda(produto, quantidade, preco):
    conn = conectar()
    cursor = conn.cursor()
    
    # Insereri uma nova venda na tabela 'vendas'
    cursor.execute('''
        INSERT INTO vendas (produto, quantidade, preco)
        VALUES (?, ?, ?)
    ''', (produto, quantidade, preco))
    
    conn.commit()
    conn.close()

def adicionar_produto_venda():
    produto = entry_produto.get()
    quantidade = entry_quantidade.get()
    preco = entry_preco.get()

    if produto and quantidade.isdigit() and preco.replace('.', '', 1).isdigit():
        quantidade = int(quantidade)
        preco = float(preco)

        # Adiciona o produto a lista de venda atual
        valor_total = quantidade * preco
        venda_atual.append((produto, quantidade, preco, valor_total))

        # Atualiza a lista exibida na interface
        lista_venda.insert("", tk.END, values=(produto, quantidade, f"R$ {preco:.2f}", f"R$ {valor_total:.2f}"))
        atualizar_valor_total()

        # Limpa os campos de entrada
        entry_produto.delete(0, tk.END)
        entry_quantidade.delete(0, tk.END)
        entry_preco.delete(0, tk.END)
    else:
        messagebox.showerror("Erro", "Preencha todos os campos corretamente.")

def atualizar_valor_total():
    total = sum(item[3] for item in venda_atual)  # Soma os valores totais dos produtos
    label_total.config(text=f"Valor Total: R$ {total:.2f}")  # Atualiza o valor total 

def finalizar_venda():
    # Aqui coleta informações sobre os produtos vendidos e o valor total
    produtos_vendidos = []
    valor_total = 0.0

    for produto, quantidade, preco, valor in venda_atual:  # Usando venda_atual
        produtos_vendidos.append(f"{produto} - R$ {preco:.2f} x {quantidade}")
        valor_total += valor  # 'valor' já é o valor total (quantidade * preço)
        registrar_venda(produto, quantidade, preco)  # Registrar a venda no banco de dados

    # Criar a mensagem
    produtos_str = "\n".join(produtos_vendidos)
    mensagem = f"Produtos Vendidos:\n{produtos_str}\n\nValor Total da Compra: R$ {valor_total:.2f}"

    # Exibir a mensagem
    messagebox.showinfo("Venda Finalizada", mensagem)

    # Limpar os campos e a lista de venda atual
    venda_atual.clear()  # Limpa a lista de venda atual
    lista_venda.delete(*lista_venda.get_children())  # Limpa a lista exibida
    label_total.config(text="Valor Total: R$ 0.00")  # Reseta o valor total na interface


def gerar_graficos():
    # Conectar ao banco de dados
    conn = conectar()  
    cursor = conn.cursor()

    # Ter as vendas do dia
    data_atual = datetime.today().strftime('%Y-%m-%d')
    cursor.execute('''
        SELECT produto, SUM(quantidade), SUM(preco)
        FROM vendas
        WHERE date(data) = ?
        GROUP BY produto
    ''', (data_atual,))
    vendas_dia = cursor.fetchall()

    # Obter total de vendas por dia 
    cursor.execute('''
        SELECT date(data) as data_venda, SUM(preco) 
        FROM vendas 
        GROUP BY date(data) 
        ORDER BY date(data) DESC
        LIMIT 7
    ''')
    vendas_por_dia = cursor.fetchall()

    # Gráfico de Barras: Produtos e suas quantidades vendidas no dia
    produtos = [venda[0] for venda in vendas_dia]
    quantidades = [venda[1] for venda in vendas_dia]

    plt.figure(figsize=(15, 5))

    # Gráfico de produtos vendidos
    plt.subplot(1, 3, 1)
    plt.bar(produtos, quantidades, color='skyblue')
    plt.title('Produtos Vendidos no Dia')
    plt.xlabel('Produtos')
    plt.ylabel('Quantidade Vendida')
    plt.xticks(rotation=90)

    # Gráfico de Linhas: Valor Total das Vendas por Dia
    datas = [venda[0] for venda in vendas_por_dia]
    valores_por_dia = [venda[1] for venda in vendas_por_dia]

    plt.subplot(1, 3, 2)
    plt.plot(datas, valores_por_dia, marker='o', color='orange')
    plt.title('Valor Total das Vendas por Dia')
    plt.xlabel('Data')
    plt.ylabel('Valor Total (R$)')
    plt.xticks(rotation=45)
    plt.grid(True)  

    # Gráfico de Barras: Valor Total das Vendas no Mês
    cursor.execute('''
        SELECT strftime('%Y-%m', data) as mes, SUM(preco) 
        FROM vendas 
        GROUP BY mes
    ''')
    vendas_mes = cursor.fetchall()

    meses = [venda[0] for venda in vendas_mes]
    valores_mes = [venda[1] for venda in vendas_mes]

    plt.subplot(1, 3, 3)
    plt.bar(meses, valores_mes, color='lightgreen')
    plt.title('Valor Total das Vendas no Mês')
    plt.xlabel('Meses')
    plt.ylabel('Valor Total (R$)')
    plt.xticks(rotation=45)

    plt.tight_layout()  # Ajusta o espaçamento 
    plt.show()  # Exibe os gráficos

    conn.close()  # Fecha a conexão após exibir os gráficos


def gerar_relatorio_vendas_dia(data):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT produto, quantidade, preco, data
        FROM vendas
        WHERE date(data) = ?
    ''', (data,))
    
    vendas = cursor.fetchall()
    conn.close()
    
    vendas_formatadas = []
    for venda in vendas:
        vendas_formatadas.append({
            'produto': venda[0],
            'quantidade': venda[1],
            'preco': venda[2],
            'data': venda[3]
        })
    
    return vendas_formatadas

def gerar_relatorio():
    data_atual = datetime.today().strftime('%Y-%m-%d')
    vendas_dia = gerar_relatorio_vendas_dia(data_atual)

    if vendas_dia:
        # Criação do DataFrame
        df = pd.DataFrame(vendas_dia)
        
        # Criação do arquivo Excel
        file_path = f'relatorio_vendas_{data_atual}.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Vendas')

            # Acessando a planilha para formatação
            ws = writer.sheets['Vendas']

            # Definindo estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alignment = Alignment(horizontal="center")

            # Aplicando estilos 
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Ajustando a largura das colunas 
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 5)  
                ws.column_dimensions[column_letter].width = adjusted_width

            # Calculando o valor total das vendas do dia
            valor_total = (df['quantidade'] * df['preco']).sum()  
            
            # Adicionando o valor total na planilha
            total_row = len(df) + 3  
            ws[f'A{total_row}'] = "Total"
            ws[f'B{total_row}'] = valor_total
            
            # Estilizando a linha total
            ws[f'A{total_row}'].font = Font(bold=True)
            ws[f'B{total_row}'].font = Font(bold=True)
            ws[f'B{total_row}'].number_format = '"R$"#,##0.00'  # Formato de moeda

            # Formatar preços dos produtos
            for index, row in df.iterrows():
                ws[f'C{index + 2}'].number_format = '"R$"#,##0.00'  # Preço unitário
                ws[f'D{index + 2}'].number_format = '"R$"#,##0.00'  # Valor total

        messagebox.showinfo("Relatório Gerado", f"Relatório gerado com sucesso: {file_path}")
    else:
        messagebox.showinfo("Relatório", "Não há vendas registradas para hoje.")



# Interface principal
root = tk.Tk()
root.title("Sistema de PDV")
root.geometry("1280x860")  
root.configure(bg='#f0f0f0')  

# Estilo do Frame
frame = tk.Frame(root, bg='#ffffff', padx=20, pady=20)
frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Título
label_titulo = tk.Label(frame, text="Sistema de PDV", font=("Arial", 24, "bold"), bg='#ffffff')
label_titulo.pack(pady=10)

# Campos de entrada
label_produto = tk.Label(frame, text="Produto", font=("Arial", 14), bg='#ffffff')
label_produto.pack()
entry_produto = tk.Entry(frame, font=("Arial", 12), width=30)
entry_produto.pack(pady=5)

label_quantidade = tk.Label(frame, text="Quantidade", font=("Arial", 14), bg='#ffffff')
label_quantidade.pack()
entry_quantidade = tk.Entry(frame, font=("Arial", 12), width=30)
entry_quantidade.pack(pady=5)

label_preco = tk.Label(frame, text="Preço", font=("Arial", 14), bg='#ffffff')
label_preco.pack()
entry_preco = tk.Entry(frame, font=("Arial", 12), width=30)
entry_preco.pack(pady=5)

# Lista de produtos da venda atual
cols = ("Produto", "Quantidade", "Preço Unitário", "Valor Total")
lista_venda = ttk.Treeview(frame, columns=cols, show='headings')
for col in cols:
    lista_venda.heading(col, text=col)
lista_venda.pack(pady=10, fill=tk.BOTH, expand=True)

# Label para valor total
label_total = tk.Label(frame, text="Valor Total: R$ 0.00", font=("Arial", 14, "bold"), bg='#ffffff')
label_total.pack(pady=5)

# Botões de ação
btn_adicionar = tk.Button(frame, text="Adicionar Produto à Venda", command=adicionar_produto_venda, font=("Arial", 12), bg='#4CAF50', fg='white')
btn_adicionar.pack(pady=10)

btn_finalizar = tk.Button(frame, text="Finalizar Venda", command=finalizar_venda, font=("Arial", 12), bg='#FF5722', fg='white')
btn_finalizar.pack(pady=10)

btn_graficos = tk.Button(frame, text="Visualizar Gráficos", command=gerar_graficos, font=("Arial", 12), bg='#2196F3', fg='white')
btn_graficos.pack(pady=10)

btn_relatorio = tk.Button(frame, text="Gerar Relatório em Planilha", command=gerar_relatorio, font=("Arial", 12), bg='#795548', fg='white')
btn_relatorio.pack(pady=10)


# Fim
root.mainloop()
